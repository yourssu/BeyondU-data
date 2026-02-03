"""Tests for extract module."""

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from src.extract.excel_reader import ExcelReader


class TestExcelReader:
    """Tests for ExcelReader class."""

    @pytest.fixture
    def sample_excel(self, tmp_path: Path) -> Path:
        """Create a sample Excel file for testing."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Add headers
        ws["A1"] = "국가"
        ws["B1"] = "대학명"
        ws["C1"] = "어학조건"

        # Add data
        ws["A2"] = "미국"
        ws["B2"] = "Harvard"
        ws["C2"] = "TOEFL 100"

        ws["A3"] = "일본"
        ws["B3"] = "Tokyo University"
        ws["C3"] = "JLPT N1"

        file_path = tmp_path / "test.xlsx"
        wb.save(file_path)
        wb.close()

        return file_path

    @pytest.fixture
    def merged_cell_excel(self, tmp_path: Path) -> Path:
        """Create an Excel file with merged cells."""
        wb = Workbook()
        ws = wb.active

        # Add headers
        ws["A1"] = "국가"
        ws["B1"] = "대학명"

        # Add data with merged cells
        ws["A2"] = "미국"
        ws["B2"] = "Harvard"
        ws["A3"] = None  # This would be part of merged cell
        ws["B3"] = "MIT"

        # Merge A2:A3
        ws.merge_cells("A2:A3")

        file_path = tmp_path / "merged.xlsx"
        wb.save(file_path)
        wb.close()

        return file_path

    def test_read_basic_excel(self, sample_excel: Path) -> None:
        """Test reading a basic Excel file."""
        reader = ExcelReader(sample_excel)
        df = reader.read()

        assert len(df) == 2
        assert "국가" in df.columns
        assert df.iloc[0]["대학명"] == "Harvard"

    def test_read_merged_cells(self, merged_cell_excel: Path) -> None:
        """Test that merged cells are properly filled."""
        reader = ExcelReader(merged_cell_excel)
        df = reader.read()

        # Both rows should have "미국" in the country column
        assert df.iloc[0]["국가"] == "미국"
        assert df.iloc[1]["국가"] == "미국"

    def test_get_sheet_names(self, sample_excel: Path) -> None:
        """Test getting sheet names."""
        reader = ExcelReader(sample_excel)
        names = reader.get_sheet_names()

        assert "Sheet1" in names

    def test_file_not_found(self, tmp_path: Path) -> None:
        """Test handling of non-existent file."""
        reader = ExcelReader(tmp_path / "nonexistent.xlsx")

        with pytest.raises(FileNotFoundError):
            reader.read()
